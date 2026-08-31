"""
Local personal tool: aggregates course-selection histograms across a set of
degree-plan .xlsx files, for a chosen semester letter.

Stdlib HTTP server + openpyxl only. Binds to 127.0.0.1 only (local-only).
"""

import json
import os
import threading
import webbrowser
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from urllib.parse import urlparse, parse_qs

import openpyxl

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
HOST = "127.0.0.1"
PORT = 8765

VALID_LETTERS = {"א", "ב", "ג", "ד", "ה", "ו"}
SUMMARY_SHEET_NAME = "סיכום"
SECTION_HEADER_TEXT = "תצוגה לפי סמסטר"


def list_xlsx_files():
    """Return sorted list of .xlsx basenames in BASE_DIR, excluding Excel lock files."""
    names = []
    for entry in os.listdir(BASE_DIR):
        full = os.path.join(BASE_DIR, entry)
        if not os.path.isfile(full):
            continue
        if entry.startswith("~$"):
            continue
        if not entry.lower().endswith(".xlsx"):
            continue
        names.append(entry)
    names.sort()
    return names


def find_summary_sheet(wb):
    """Return the worksheet to use, or None if none found."""
    if SUMMARY_SHEET_NAME in wb.sheetnames:
        return wb[SUMMARY_SHEET_NAME]
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == SECTION_HEADER_TEXT:
                    return ws
    return None


def extract_courses_for_semester(full_path, semester):
    """
    Open the workbook fresh, find the summary sheet, and extract the set of
    unique courses (by code, falling back to name) tagged with the given
    semester letter. Returns a dict: key -> {"code": ..., "name": ...}
    """
    wb = openpyxl.load_workbook(full_path, data_only=True, read_only=True)
    try:
        ws = find_summary_sheet(wb)
        if ws is None:
            raise ValueError("no summary sheet found")

        found = {}
        for row in ws.iter_rows():
            for i, cell in enumerate(row):
                if i < 3:
                    continue
                value = cell.value
                if not isinstance(value, str):
                    continue
                if value.strip() != semester:
                    continue
                name = row[i - 2].value
                if not isinstance(name, str) or not name.strip():
                    continue
                code = row[i - 3].value
                key = code if code is not None else name.strip()
                if key not in found:
                    found[key] = {"code": code, "name": name.strip()}
        return found
    finally:
        try:
            wb.close()
        except Exception:
            pass


class Handler(BaseHTTPRequestHandler):
    def log_message(self, format, *args):
        # Silence default noisy request logging.
        pass

    def _send_json(self, status, payload):
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _send_html_file(self, filename):
        full_path = os.path.join(BASE_DIR, filename)
        try:
            with open(full_path, "rb") as f:
                body = f.read()
        except OSError as e:
            self._send_json(500, {"error": "could not read %s: %s" % (filename, str(e))})
            return
        self.send_response(200)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _handle_files(self):
        names = list_xlsx_files()
        self._send_json(200, names)

    def _handle_histogram(self, query):
        semester_list = query.get("semester", [])
        semester = semester_list[0].strip() if semester_list else ""
        if semester not in VALID_LETTERS:
            self._send_json(400, {"error": "missing or invalid 'semester' parameter; must be one of %s" % ", ".join(sorted(VALID_LETTERS))})
            return

        requested_files = query.get("files", [])
        existing_files = set(list_xlsx_files())

        errors = []
        aggregate = {}  # key -> {"code":..., "name":..., "count": int}

        for name in requested_files:
            try:
                if os.path.basename(name) != name:
                    errors.append({"file": name, "message": "invalid file name"})
                    continue
                if not name.endswith(".xlsx"):
                    errors.append({"file": name, "message": "not an .xlsx file"})
                    continue
                if name.startswith("~$"):
                    errors.append({"file": name, "message": "invalid file name"})
                    continue
                if name not in existing_files:
                    errors.append({"file": name, "message": "file not found"})
                    continue

                full_path = os.path.join(BASE_DIR, name)
                courses = extract_courses_for_semester(full_path, semester)
                for key, info in courses.items():
                    entry = aggregate.get(key)
                    if entry is None:
                        aggregate[key] = {"code": info["code"], "name": info["name"], "count": 1}
                    else:
                        entry["count"] += 1
            except Exception as e:
                errors.append({"file": name, "message": str(e)})

        results = list(aggregate.values())
        results.sort(key=lambda r: (-r["count"], r["name"]))

        self._send_json(200, {"semester": semester, "results": results, "errors": errors})

    def do_GET(self):
        try:
            parsed = urlparse(self.path)
            path = parsed.path

            if path == "/":
                self._send_html_file("index.html")
            elif path == "/api/files":
                self._handle_files()
            elif path == "/api/histogram":
                query = parse_qs(parsed.query)
                self._handle_histogram(query)
            else:
                self._send_json(404, {"error": "not found"})
        except Exception as e:
            try:
                self._send_json(500, {"error": str(e)})
            except Exception:
                pass


def main():
    server = ThreadingHTTPServer((HOST, PORT), Handler)
    url = "http://%s:%d/" % (HOST, PORT)

    def open_browser():
        webbrowser.open(url)

    threading.Timer(0.5, open_browser).start()

    print("Serving on %s (base dir: %s)" % (url, BASE_DIR))
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        pass
    finally:
        server.server_close()


if __name__ == "__main__":
    main()
