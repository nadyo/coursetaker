"""
Permanent, rerunnable, end-to-end test suite for server.py.

This does NOT import server.py's internal functions. Every test starts the
real server.py as a subprocess and talks to it over real HTTP (stdlib
urllib/http.client only), exactly the way a browser would. Fixtures are
built with openpyxl into temp directories; the real project directory
(with the user's real .xlsx) is only ever touched with GET requests.

Run with:
    py tests\\test_server.py
or:
    py -m unittest tests.test_server
"""

import concurrent.futures
import http.client
import json
import os
import shutil
import socket
import subprocess
import sys
import tempfile
import time
import unittest
import urllib.error
import urllib.parse
import urllib.request

import openpyxl

# ---------------------------------------------------------------------------
# Paths / constants
# ---------------------------------------------------------------------------

TESTS_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(TESTS_DIR)
SERVER_SRC_PATH = os.path.join(PROJECT_DIR, "server.py")
INDEX_HTML_PATH = os.path.join(PROJECT_DIR, "index.html")

REAL_XLSX_NAME = "מבנה תואר - ישראל ישראלי.xlsx"

# Mirrors the documented xlsx contract in docs/ARCHITECTURE.md. These are
# plain string literals describing the *file format*, not imports of
# server.py's internals.
SUMMARY_SHEET_NAME = "סיכום"
SECTION_HEADER_TEXT = "תצוגה לפי סמסטר"
VALID_LETTERS = ("א", "ב", "ג", "ד", "ה", "ו")

REAL_DIR_HARNESS_SCRIPT_NAME = "_takecourses_test_harness_server.py"

# Verified directly against the real fixture file two independent ways:
#   1) calling server.py's own extract_courses_for_semester() in-process
#   2) a raw openpyxl cell dump of the "סיכום" sheet
# Both agree on this set. See test_01's docstring for a note about a
# discrepancy with an earlier assumption about one of the course names.
REAL_FILE_SEMESTER_A_COURSES = {
    (77101, "מכניקה ויחסות פרטית"),
    (77110, "מעבדה בפיסיקה א' תלפיות"),
    (67109, "מבני נתונים"),
    (67925, "מ-NAND ל-TETRIS"),
    (80114, "מתמטיקה שימושית 1"),
    (67125, "מבוא לתכנות מונחה עצמים"),
}


# ---------------------------------------------------------------------------
# Server process management
# ---------------------------------------------------------------------------

def _free_port():
    s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    s.bind(("127.0.0.1", 0))
    port = s.getsockname()[1]
    s.close()
    return port


def _read_server_source():
    with open(SERVER_SRC_PATH, "r", encoding="utf-8") as f:
        return f.read()


def _patched_server_source(port):
    """
    Returns server.py's source with PORT patched to an isolated free port
    and the on-startup webbrowser.open() call disabled (so running this
    suite doesn't pop open a browser tab per server start).
    """
    src = _read_server_source()

    port_marker = "PORT = 8765"
    if port_marker not in src:
        raise RuntimeError(
            "test harness could not find %r in server.py to patch the port "
            "-- server.py may have changed; update the test harness." % port_marker
        )
    src = src.replace(port_marker, "PORT = %d" % port, 1)

    browser_marker = "webbrowser.open(url)"
    if browser_marker not in src:
        raise RuntimeError(
            "test harness could not find %r in server.py to disable the "
            "auto-opened browser tab; update the test harness." % browser_marker
        )
    src = src.replace(browser_marker, "pass  # disabled by test harness", 1)

    return src


def _wait_for_server(port, timeout=10.0):
    deadline = time.time() + timeout
    last_err = None
    while time.time() < deadline:
        try:
            conn = http.client.HTTPConnection("127.0.0.1", port, timeout=1)
            try:
                conn.request("GET", "/api/files")
                resp = conn.getresponse()
                resp.read()
            finally:
                conn.close()
            return
        except Exception as e:
            last_err = e
            time.sleep(0.1)
    raise RuntimeError("server on port %d never became ready: %r" % (port, last_err))


class ManagedServer(object):
    """
    Drops a patched copy of server.py (isolated port, no auto browser-open)
    into base_dir, launches it as a real subprocess, and waits for it to
    accept connections. stop() always kills the process and removes the
    dropped script, even if start() only partially succeeded.
    """

    def __init__(self, base_dir, script_name="server.py"):
        self.base_dir = base_dir
        self.script_name = script_name
        self.port = None
        self.proc = None
        self.script_path = os.path.join(base_dir, script_name)

    def start(self):
        self.port = _free_port()
        patched = _patched_server_source(self.port)
        with open(self.script_path, "w", encoding="utf-8") as f:
            f.write(patched)

        target_index = os.path.join(self.base_dir, "index.html")
        if not os.path.exists(target_index):
            shutil.copyfile(INDEX_HTML_PATH, target_index)

        self.proc = subprocess.Popen(
            [sys.executable, self.script_path],
            cwd=self.base_dir,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
        )
        try:
            _wait_for_server(self.port)
        except Exception:
            self.stop()
            raise

    def stop(self):
        if self.proc is not None:
            try:
                self.proc.terminate()
                try:
                    self.proc.wait(timeout=5)
                except subprocess.TimeoutExpired:
                    self.proc.kill()
                    self.proc.wait(timeout=5)
            except Exception:
                pass
            finally:
                self.proc = None
        if self.script_path and os.path.exists(self.script_path):
            try:
                os.remove(self.script_path)
            except OSError:
                pass

    def url(self, path):
        return "http://127.0.0.1:%d%s" % (self.port, path)


# ---------------------------------------------------------------------------
# HTTP helpers
# ---------------------------------------------------------------------------

def http_get(url):
    """GET url. Returns (status, headers, raw_bytes) -- never raises on 4xx/5xx."""
    try:
        resp = urllib.request.urlopen(url, timeout=10)
        try:
            return resp.status, resp.headers, resp.read()
        finally:
            resp.close()
    except urllib.error.HTTPError as e:
        body = e.read()
        return e.code, e.headers, body


def http_get_json(url):
    status, headers, raw = http_get(url)
    return status, headers, json.loads(raw.decode("utf-8")), raw


def build_url(server, path, params):
    qs = urllib.parse.urlencode(params, doseq=True)
    return server.url(path) + ("?" + qs if qs else "")


# ---------------------------------------------------------------------------
# Fixture builders (openpyxl)
# ---------------------------------------------------------------------------

def _new_workbook(sheet_title):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    return wb, ws


def write_pattern_workbook(path, rows, sheet_name=SUMMARY_SHEET_NAME, start_row=1):
    """
    rows: list of (code, name, credits, letter) tuples. Each tuple is
    written to one row at columns A-D (code=A, name=B, credits=C,
    letter=D), matching the documented 4-cell left-to-right pattern with
    the letter landing at column index 3 (>= 3, so it's a legal scan
    position for extraction).
    """
    wb, ws = _new_workbook(sheet_name)
    for offset, (code, name, credits, letter) in enumerate(rows):
        r = start_row + offset
        ws.cell(row=r, column=1, value=code)
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=credits)
        ws.cell(row=r, column=4, value=letter)
    wb.save(path)


def write_fallback_workbook(path, header_cell, data_row):
    """
    A workbook with NO sheet literally named "סיכום", but some other sheet
    has a cell containing SECTION_HEADER_TEXT plus a valid data row
    elsewhere on that same sheet.
    """
    wb, ws = _new_workbook("Sheet2")
    hr, hc = header_cell
    ws.cell(row=hr, column=hc, value=SECTION_HEADER_TEXT)
    r = data_row["row"]
    ws.cell(row=r, column=1, value=data_row["code"])
    ws.cell(row=r, column=2, value=data_row["name"])
    ws.cell(row=r, column=3, value=data_row["credits"])
    ws.cell(row=r, column=4, value=data_row["letter"])
    wb.save(path)


def write_no_summary_workbook(path):
    """Neither a "סיכום" sheet nor any cell containing SECTION_HEADER_TEXT."""
    wb, ws = _new_workbook("Random")
    ws.cell(row=1, column=1, value="nothing relevant here")
    wb.save(path)


def write_merged_header_workbook(path, valid_row, merged_row):
    """
    One normal valid data row, plus a row where the "name" cell (column B,
    the position 2 cells left of the letter) is the non-anchor half of a
    merged range -- so reading it directly returns None, exactly how a
    real header/total row in the actual template reads.
    """
    wb, ws = _new_workbook(SUMMARY_SHEET_NAME)
    vr = valid_row["row"]
    ws.cell(row=vr, column=1, value=valid_row["code"])
    ws.cell(row=vr, column=2, value=valid_row["name"])
    ws.cell(row=vr, column=3, value=valid_row["credits"])
    ws.cell(row=vr, column=4, value=valid_row["letter"])

    mr = merged_row["row"]
    ws.cell(row=mr, column=1, value=merged_row["anchor_text"])
    ws.merge_cells(start_row=mr, start_column=1, end_row=mr, end_column=2)
    ws.cell(row=mr, column=3, value=0)
    ws.cell(row=mr, column=4, value=merged_row["letter"])
    wb.save(path)


def write_corrupt_file(path):
    """Not a valid xlsx at all -- plain bytes."""
    with open(path, "wb") as f:
        f.write(b"this is not a zip file / not a real xlsx, just plain text bytes 12345")


def write_minimal_valid_workbook(path):
    wb, ws = _new_workbook(SUMMARY_SHEET_NAME)
    ws.cell(row=1, column=1, value="placeholder")
    wb.save(path)


def write_lock_placeholder(path):
    """Excel lock-file-named placeholder. Content is irrelevant -- it must
    never be opened by the server."""
    with open(path, "wb") as f:
        f.write(b"not real xlsx content, must never be read")


# ---------------------------------------------------------------------------
# Test 1: extraction correctness against the REAL example file
# ---------------------------------------------------------------------------

class RealDirectoryTests(unittest.TestCase):
    """Points a server instance at the real project directory (read-only:
    GET requests only, never modifies/deletes the real .xlsx)."""

    @classmethod
    def setUpClass(cls):
        harness_path = os.path.join(PROJECT_DIR, REAL_DIR_HARNESS_SCRIPT_NAME)
        if os.path.exists(harness_path):
            # Leftover from a previous crashed run -- clear it before we start.
            os.remove(harness_path)

        cls.server = ManagedServer(PROJECT_DIR, script_name=REAL_DIR_HARNESS_SCRIPT_NAME)
        cls.server.start()

    @classmethod
    def tearDownClass(cls):
        cls.server.stop()

    def test_01_extraction_matches_real_file(self):
        """
        NOTE on a discrepancy vs. an earlier assumption: the course list
        originally expected here included "מבוא למדעי המחשב" (general
        Intro to CS). Direct inspection of the real file (both by calling
        server.py's own extraction function in-process, and by dumping raw
        cells of the "סיכום" sheet) shows that name is NOT tagged with a
        semester letter -- it sits in the file's separate exemptions
        block, tagged "זיכוי" (credit/exemption), which is correctly
        excluded because "זיכוי" is not in {א,ב,ג,ד,ה,ו}. The actual 6th
        semester-א course in the file today is "מבוא לתכנות מונחה עצמים"
        (Intro to OOP, code 67125). The file's mtime is very recent, so
        this is most likely because the user has since edited their real
        degree-plan file, not a bug in server.py. This test asserts
        against the verified-current, verified-correct data so it's a
        real regression guard going forward; see the final report for
        this flagged prominently.
        """
        url = build_url(
            self.server,
            "/api/histogram",
            {"semester": "א", "files": REAL_XLSX_NAME},
        )
        status, headers, body, raw = http_get_json(url)

        self.assertEqual(status, 200, "expected HTTP 200, got %d; body=%r" % (status, body))
        self.assertEqual(body.get("errors"), [], "expected no errors, got %r" % (body.get("errors"),))

        results = body.get("results", [])
        self.assertEqual(
            len(results), 6,
            "expected exactly 6 courses for semester א in the real file, got %d: %r" % (len(results), results),
        )
        found = {(r["code"], r["name"]) for r in results}
        self.assertEqual(
            found, REAL_FILE_SEMESTER_A_COURSES,
            "extracted course set does not match verified-correct set.\n  extracted: %r\n  expected:  %r"
            % (found, REAL_FILE_SEMESTER_A_COURSES),
        )
        for r in results:
            self.assertEqual(
                r["count"], 1,
                "expected count 1 for %r (single file requested), got %d" % (r["name"], r["count"]),
            )


# ---------------------------------------------------------------------------
# Tests 2-17: synthetic fixtures in an isolated temp directory
# ---------------------------------------------------------------------------

# Filenames used by the synthetic fixture directory.
F02_ZERO = "test02_zero_courses.xlsx"
F03_FALLBACK = "test03_fallback.xlsx"
F04_NO_SUMMARY = "test04_no_summary.xlsx"
F05_DUP_INFILE = "test05_infile_duplicate.xlsx"
F06_FALSY_CODE = "test06_falsy_code.xlsx"
F07_MERGED = "test07_merged_header.xlsx"
F08_CORRUPT = "corrupt.xlsx"
F08_VALID = "test08_valid_alongside_corrupt.xlsx"
F09_A = "test09_multi_a.xlsx"
F09_B = "test09_multi_b.xlsx"
F09_C = "test09_multi_c.xlsx"
F10_REAL = "test10_real.xlsx"
F10_LOCK = "~$test10_real.xlsx"
F12_VALID = "test12_valid.xlsx"
F13_VALID = "test13_valid.xlsx"
F14_DUP_NAME = "test14_dup_filename.xlsx"
F16_HEADERS = "test16_headers.xlsx"


def _build_synthetic_fixtures(base_dir):
    # --- Test 2: valid header-pattern position for semester א, zero data rows.
    write_pattern_workbook(os.path.join(base_dir, F02_ZERO), rows=[(None, None, None, "א")])

    # --- Test 3: fallback sheet detection.
    write_fallback_workbook(
        os.path.join(base_dir, F03_FALLBACK),
        header_cell=(1, 1),
        data_row={"row": 5, "code": 90001, "name": "קורס פולבק", "credits": 3, "letter": "א"},
    )

    # --- Test 4: no summary sheet and no fallback header text anywhere.
    write_no_summary_workbook(os.path.join(base_dir, F04_NO_SUMMARY))

    # --- Test 5: same course code appears twice in the same file/semester.
    write_pattern_workbook(
        os.path.join(base_dir, F05_DUP_INFILE),
        rows=[
            (555001, "קורס כפול בקובץ", 3, "א"),
            (555002, "קורס אחר", 2, "ב"),
            (555001, "קורס כפול בקובץ", 3, "א"),
        ],
    )

    # --- Test 6: falsy/missing code (None, and 0).
    write_pattern_workbook(
        os.path.join(base_dir, F06_FALSY_CODE),
        rows=[
            (None, "קורס בלי קוד", 2, "א"),
            (0, "קורס עם קוד אפס", 2, "א"),
        ],
    )

    # --- Test 7: merged-cell header row must be silently skipped.
    write_merged_header_workbook(
        os.path.join(base_dir, F07_MERGED),
        valid_row={"row": 1, "code": 700001, "name": "קורס תקין אחרי מיזוג", "credits": 3, "letter": "א"},
        merged_row={"row": 2, "anchor_text": "כותרת ממוזגת", "letter": "א"},
    )

    # --- Test 8: a corrupted xlsx alongside a valid one.
    write_corrupt_file(os.path.join(base_dir, F08_CORRUPT))
    write_pattern_workbook(
        os.path.join(base_dir, F08_VALID),
        rows=[(800001, "קורס תקין ליד קובץ פגום", 3, "א")],
    )

    # --- Test 9: multi-file aggregation (X in all 3, Y in 2, Z in 1).
    write_pattern_workbook(
        os.path.join(base_dir, F09_A),
        rows=[(100, "קורס X", 3, "א"), (200, "קורס Y", 3, "א")],
    )
    write_pattern_workbook(
        os.path.join(base_dir, F09_B),
        rows=[(100, "קורס X", 3, "א"), (200, "קורס Y", 3, "א")],
    )
    write_pattern_workbook(
        os.path.join(base_dir, F09_C),
        rows=[(100, "קורס X", 3, "א"), (300, "קורס Z", 3, "א")],
    )

    # --- Test 10: /api/files must exclude Excel lock files.
    write_minimal_valid_workbook(os.path.join(base_dir, F10_REAL))
    write_lock_placeholder(os.path.join(base_dir, F10_LOCK))

    # --- Test 12: a valid file, for combining with a nonexistent filename.
    write_pattern_workbook(
        os.path.join(base_dir, F12_VALID),
        rows=[(120001, "קורס לבדיקת קובץ לא קיים", 4, "ד")],
    )

    # --- Test 13: a valid file, for combining with path-traversal attempts.
    write_pattern_workbook(
        os.path.join(base_dir, F13_VALID),
        rows=[(130001, "קורס לבדיקת מניעת מעקף נתיב", 2, "א")],
    )

    # --- Test 14: a valid file, requested twice via duplicate files= params.
    write_pattern_workbook(
        os.path.join(base_dir, F14_DUP_NAME),
        rows=[(140001, "קורס לבדיקת כפילות בשם קובץ", 2, "ו")],
    )

    # --- Test 16: Hebrew-bearing response, for Content-Type/Length checks.
    write_pattern_workbook(
        os.path.join(base_dir, F16_HEADERS),
        rows=[(160001, "קורס עם תווים בעברית לבדיקת קידוד תוכן", 3, "ב")],
    )


class SyntheticServerTests(unittest.TestCase):
    """One shared temp directory + one shared server instance for all of
    tests 2-17 (each test uses its own uniquely-named fixture file(s), so
    they don't interfere with each other)."""

    @classmethod
    def setUpClass(cls):
        cls.tmpdir = tempfile.mkdtemp(prefix="takecourses_test_")
        try:
            _build_synthetic_fixtures(cls.tmpdir)
            cls.server = ManagedServer(cls.tmpdir, script_name="server.py")
            cls.server.start()
        except Exception:
            shutil.rmtree(cls.tmpdir, ignore_errors=True)
            raise

    @classmethod
    def tearDownClass(cls):
        cls.server.stop()
        shutil.rmtree(cls.tmpdir, ignore_errors=True)

    # -- Test 2 --------------------------------------------------------

    def test_02_zero_course_semester_block(self):
        url = build_url(self.server, "/api/histogram", {"semester": "א", "files": F02_ZERO})
        status, _headers, body, _raw = http_get_json(url)
        self.assertEqual(status, 200)
        self.assertEqual(body.get("results"), [], "expected zero results, got %r" % (body.get("results"),))
        self.assertEqual(body.get("errors"), [], "expected zero errors, got %r" % (body.get("errors"),))

    # -- Test 3 --------------------------------------------------------

    def test_03_fallback_sheet_detection(self):
        url = build_url(self.server, "/api/histogram", {"semester": "א", "files": F03_FALLBACK})
        status, _headers, body, _raw = http_get_json(url)
        self.assertEqual(status, 200)
        self.assertEqual(body.get("errors"), [], "fallback extraction should not error: %r" % (body.get("errors"),))
        results = body.get("results", [])
        self.assertEqual(len(results), 1, "expected exactly 1 course via fallback, got %r" % (results,))
        self.assertEqual(results[0]["name"], "קורס פולבק")
        self.assertEqual(results[0]["count"], 1)

    # -- Test 4 --------------------------------------------------------

    def test_04_no_summary_sheet_at_all(self):
        url = build_url(self.server, "/api/histogram", {"semester": "א", "files": F04_NO_SUMMARY})
        status, _headers, body, _raw = http_get_json(url)
        self.assertEqual(status, 200, "should not crash the whole request, got HTTP %d" % status)
        self.assertEqual(body.get("results"), [], "should contribute no results")
        errors = body.get("errors", [])
        self.assertEqual(len(errors), 1, "expected exactly one error entry, got %r" % (errors,))
        self.assertEqual(errors[0]["file"], F04_NO_SUMMARY)
        self.assertTrue(errors[0].get("message"), "error entry should carry a message")

    # -- Test 5 --------------------------------------------------------

    def test_05_infile_duplicate_course_counts_once(self):
        url = build_url(self.server, "/api/histogram", {"semester": "א", "files": F05_DUP_INFILE})
        status, _headers, body, _raw = http_get_json(url)
        self.assertEqual(status, 200)
        self.assertEqual(body.get("errors"), [])
        results = body.get("results", [])
        self.assertEqual(len(results), 1, "expected exactly 1 unique course, got %r" % (results,))
        self.assertEqual(results[0]["name"], "קורס כפול בקובץ")
        self.assertEqual(
            results[0]["count"], 1,
            "a course duplicated within one file's block must still count 1 for that file, got %d"
            % results[0]["count"],
        )

    # -- Test 6 --------------------------------------------------------

    def test_06_falsy_or_missing_code(self):
        url = build_url(self.server, "/api/histogram", {"semester": "א", "files": F06_FALSY_CODE})
        status, _headers, body, _raw = http_get_json(url)
        self.assertEqual(status, 200)
        self.assertEqual(body.get("errors"), [])
        results = body.get("results", [])
        by_name = {r["name"]: r for r in results}

        self.assertIn("קורס בלי קוד", by_name, "course keyed by name (code=None) missing: %r" % (results,))
        self.assertIsNone(by_name["קורס בלי קוד"]["code"])
        self.assertEqual(by_name["קורס בלי קוד"]["count"], 1)

        self.assertIn("קורס עם קוד אפס", by_name, "course with code=0 missing: %r" % (results,))
        self.assertEqual(by_name["קורס עם קוד אפס"]["code"], 0)
        self.assertEqual(by_name["קורס עם קוד אפס"]["count"], 1)

        self.assertEqual(len(results), 2, "expected exactly 2 courses, got %r" % (results,))

    # -- Test 7 --------------------------------------------------------

    def test_07_merged_cell_header_row_silently_skipped(self):
        url = build_url(self.server, "/api/histogram", {"semester": "א", "files": F07_MERGED})
        status, _headers, body, _raw = http_get_json(url)
        self.assertEqual(status, 200)
        self.assertEqual(body.get("errors"), [], "merged-header row must not cause an error: %r" % (body.get("errors"),))
        results = body.get("results", [])
        self.assertEqual(len(results), 1, "expected only the one real course, got %r" % (results,))
        self.assertEqual(results[0]["name"], "קורס תקין אחרי מיזוג")
        self.assertEqual(results[0]["count"], 1)

    # -- Test 8 --------------------------------------------------------

    def test_08_corrupted_file_reported_as_error_valid_file_still_works(self):
        url = build_url(self.server, "/api/histogram", {"semester": "א", "files": [F08_CORRUPT, F08_VALID]})
        status, _headers, body, _raw = http_get_json(url)
        self.assertEqual(status, 200, "corrupted file must not break the whole request")

        errors = body.get("errors", [])
        self.assertEqual(len(errors), 1, "expected exactly one error (the corrupt file), got %r" % (errors,))
        self.assertEqual(errors[0]["file"], F08_CORRUPT)
        self.assertTrue(errors[0].get("message"), "corrupt-file error should carry a message")

        results = body.get("results", [])
        self.assertEqual(len(results), 1, "valid file's course should still come back: %r" % (results,))
        self.assertEqual(results[0]["name"], "קורס תקין ליד קובץ פגום")
        self.assertEqual(results[0]["count"], 1)

    # -- Test 9 --------------------------------------------------------

    def test_09_multi_file_aggregation_counts(self):
        url = build_url(self.server, "/api/histogram", {"semester": "א", "files": [F09_A, F09_B, F09_C]})
        status, _headers, body, _raw = http_get_json(url)
        self.assertEqual(status, 200)
        self.assertEqual(body.get("errors"), [])
        results = body.get("results", [])
        by_name = {r["name"]: r["count"] for r in results}

        self.assertEqual(by_name.get("קורס X"), 3, "course X should appear in all 3 files: %r" % (by_name,))
        self.assertEqual(by_name.get("קורס Y"), 2, "course Y should appear in 2 files: %r" % (by_name,))
        self.assertEqual(by_name.get("קורס Z"), 1, "course Z should appear in 1 file: %r" % (by_name,))
        self.assertEqual(len(results), 3, "expected exactly 3 distinct courses, got %r" % (results,))

    # -- Test 10 --------------------------------------------------------

    def test_10_files_endpoint_excludes_lock_files(self):
        url = self.server.url("/api/files")
        status, _headers, body, _raw = http_get_json(url)
        self.assertEqual(status, 200)
        self.assertIn(F10_REAL, body, "real fixture file should be listed: %r" % (body,))
        self.assertNotIn(F10_LOCK, body, "Excel lock file must be excluded from listing: %r" % (body,))
        for name in body:
            self.assertFalse(name.startswith("~$"), "no listed file should start with ~$: %r" % (name,))

    # -- Test 11 --------------------------------------------------------

    def test_11_invalid_semester_handling(self):
        # missing 'semester' entirely
        url = build_url(self.server, "/api/histogram", {"files": F12_VALID})
        status, _headers, body, _raw = http_get_json(url)
        self.assertEqual(status, 400, "missing semester should be HTTP 400, got %d" % status)
        self.assertIn("error", body, "missing-semester response should carry an 'error' key: %r" % (body,))

        # empty string semester
        url = build_url(self.server, "/api/histogram", {"semester": "", "files": F12_VALID})
        status, _headers, body, _raw = http_get_json(url)
        self.assertEqual(status, 400, "empty semester should be HTTP 400, got %d" % status)
        self.assertIn("error", body, "empty-semester response should carry an 'error' key: %r" % (body,))

        # invalid value (not one of א/ב/ג/ד/ה/ו)
        url = build_url(self.server, "/api/histogram", {"semester": "X", "files": F12_VALID})
        status, _headers, body, _raw = http_get_json(url)
        self.assertEqual(status, 400, "invalid semester 'X' should be HTTP 400, got %d" % status)
        self.assertIn("error", body, "invalid-semester response should carry an 'error' key: %r" % (body,))

    # -- Test 12 --------------------------------------------------------

    def test_12_nonexistent_file_in_request(self):
        url = build_url(
            self.server,
            "/api/histogram",
            {"semester": "ד", "files": [F12_VALID, "does_not_exist_xyz.xlsx"]},
        )
        status, _headers, body, _raw = http_get_json(url)
        self.assertEqual(status, 200, "a missing requested file must not break the whole request")

        errors = body.get("errors", [])
        self.assertEqual(len(errors), 1, "expected exactly one error, got %r" % (errors,))
        self.assertEqual(errors[0]["file"], "does_not_exist_xyz.xlsx")
        self.assertEqual(errors[0]["message"], "file not found")

        results = body.get("results", [])
        self.assertEqual(len(results), 1, "the other valid file's results should still come back: %r" % (results,))
        self.assertEqual(results[0]["name"], "קורס לבדיקת קובץ לא קיים")
        self.assertEqual(results[0]["count"], 1)

    # -- Test 13 --------------------------------------------------------

    def test_13_path_traversal_rejected(self):
        windows_style = "..\\..\\Windows\\win.ini"
        unix_style = "../../etc/passwd"
        absolute_style = "C:\\Windows\\win.ini"

        url = build_url(
            self.server,
            "/api/histogram",
            {"semester": "א", "files": [windows_style, unix_style, absolute_style, F13_VALID]},
        )
        status, _headers, body, _raw = http_get_json(url)
        self.assertEqual(status, 200)

        errors_by_file = {e["file"]: e["message"] for e in body.get("errors", [])}
        for attempt in (windows_style, unix_style, absolute_style):
            self.assertIn(
                attempt, errors_by_file,
                "path traversal attempt %r should be rejected via errors, got errors=%r"
                % (attempt, body.get("errors")),
            )
            self.assertEqual(
                errors_by_file[attempt], "invalid file name",
                "expected 'invalid file name' for %r, got %r" % (attempt, errors_by_file[attempt]),
            )
        self.assertEqual(len(body.get("errors", [])), 3, "no unexpected extra errors: %r" % (body.get("errors"),))

        results = body.get("results", [])
        self.assertEqual(len(results), 1, "the valid file should still be processed: %r" % (results,))
        self.assertEqual(results[0]["name"], "קורס לבדיקת מניעת מעקף נתיב")

    # -- Test 14 --------------------------------------------------------

    def test_14_duplicate_filename_in_query_characterization(self):
        url = build_url(
            self.server,
            "/api/histogram",
            {"semester": "ו", "files": [F14_DUP_NAME, F14_DUP_NAME]},
        )
        status, _headers, body, _raw = http_get_json(url)
        self.assertEqual(status, 200)
        self.assertEqual(body.get("errors"), [])
        results = body.get("results", [])
        self.assertEqual(len(results), 1)
        # NOTE: server.py does not dedupe the files= list itself; the
        # 'requested_files' loop in _handle_histogram runs once per
        # occurrence of a filename in the query string, and each pass
        # increments the aggregate count independently. So passing the
        # same filename twice currently DOES double-count that file's
        # courses (observed: count == 2, not 1). This is a known,
        # currently-unreachable-via-the-UI edge case (the UI's checkboxes
        # can't produce a duplicate 'files' param) -- documented here as a
        # characterization test, not a silent assumption.
        self.assertEqual(
            results[0]["count"], 2,
            "characterization: duplicate files= currently double-counts (observed count=%d)"
            % results[0]["count"],
        )

    # -- Test 15 --------------------------------------------------------

    def test_15_unknown_route_404(self):
        status, _headers, body, _raw = http_get_json(self.server.url("/api/nonexistent"))
        self.assertEqual(status, 404, "unknown route should be HTTP 404, got %d" % status)

    # -- Test 16 --------------------------------------------------------

    def test_16_content_type_and_length_for_hebrew_body(self):
        url = build_url(self.server, "/api/histogram", {"semester": "ב", "files": F16_HEADERS})
        status, headers, raw = http_get(url)
        self.assertEqual(status, 200)

        content_type = headers.get("Content-Type")
        self.assertEqual(content_type, "application/json; charset=utf-8", "unexpected Content-Type: %r" % content_type)

        content_length_header = headers.get("Content-Length")
        self.assertIsNotNone(content_length_header, "response should carry a Content-Length header")
        self.assertEqual(
            int(content_length_header), len(raw),
            "Content-Length header (%s) must equal actual UTF-8 byte length (%d)"
            % (content_length_header, len(raw)),
        )

        text = raw.decode("utf-8")
        self.assertGreater(
            len(raw), len(text),
            "response contains Hebrew text, so UTF-8 byte length must exceed character "
            "count (Hebrew chars are 2 bytes each) -- got byte_len=%d, char_len=%d"
            % (len(raw), len(text)),
        )
        self.assertIn("קורס עם תווים בעברית", text, "sanity check: Hebrew course name should be in the body")

    # -- Test 17 --------------------------------------------------------

    def test_17_concurrent_requests_do_not_crash_server(self):
        files_url = self.server.url("/api/files")
        histogram_url = build_url(self.server, "/api/histogram", {"semester": "א", "files": F13_VALID})
        urls = ([files_url] * 5) + ([histogram_url] * 5)

        timings = []

        def do_request(u):
            t0 = time.time()
            status, _headers, raw = http_get(u)
            t1 = time.time()
            return status, raw, t0, t1

        with concurrent.futures.ThreadPoolExecutor(max_workers=len(urls)) as pool:
            outcomes = list(pool.map(do_request, urls))

        for status, raw, t0, t1 in outcomes:
            self.assertEqual(status, 200, "every concurrent request should succeed with HTTP 200")
            json.loads(raw.decode("utf-8"))  # must be valid JSON, not a truncated/garbled response
            timings.append((t0, t1))

        # Informational only (not asserted): whether client-observed request
        # windows overlap depends on whether server.py's HTTP server class
        # handles connections one at a time (stdlib HTTPServer) or on a
        # thread per request (ThreadingHTTPServer). Either way this test only
        # asserts that every request completes cleanly with a valid response
        # -- it does not assert serialization or concurrency either way,
        # since that's an implementation detail of server.py that may change.
        overlapping = any(
            a_start < b_end and b_start < a_end
            for i, (a_start, a_end) in enumerate(timings)
            for j, (b_start, b_end) in enumerate(timings)
            if i != j
        )
        print(
            "    [test_17 info] requests appeared to %s"
            % ("overlap/run concurrently" if overlapping else "serialize")
        )


# ---------------------------------------------------------------------------
# Runner with a clear PASS/FAIL summary
# ---------------------------------------------------------------------------

def build_suite():
    loader = unittest.TestLoader()
    suite = unittest.TestSuite()
    suite.addTests(loader.loadTestsFromTestCase(RealDirectoryTests))
    suite.addTests(loader.loadTestsFromTestCase(SyntheticServerTests))
    return suite


def main():
    suite = build_suite()
    runner = unittest.TextTestRunner(verbosity=2)
    result = runner.run(suite)

    total = result.testsRun
    failed = list(result.failures) + list(result.errors)
    passed = total - len(failed)

    print()
    print("=" * 70)
    print("SUMMARY: %d/%d passed" % (passed, total))
    if failed:
        print("FAILED TESTS:")
        for test, tb in failed:
            tb_lines = tb.strip().splitlines()
            reason = tb_lines[-1] if tb_lines else "(no traceback)"
            print("  - %s: %s" % (test.id(), reason))
    print("=" * 70)

    sys.exit(0 if not failed else 1)


if __name__ == "__main__":
    main()
